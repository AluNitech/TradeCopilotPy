from config import DATA_HANDLER_OPTIONS
from trade_copilot_py.data_handler import get_multi_timeframe_data, fetch_daytrading_morning_ranking
from trade_copilot_py.strategy import advise_from_data

import os
import pandas as pd
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime


def main() -> None:
    ranking = fetch_daytrading_morning_ranking()
    if not ranking:
        print("ランキングが空でした。終了します。")
        return

    opts = DATA_HANDLER_OPTIONS

    rows = []  # Excel出力用の集計

    def process_one(stock):
        ticker = stock.code + ".T"
        print(
            f"\n=== 銘柄: [{stock.code}] {stock.name} ({stock.market}) / ティッカー={ticker} ===\n"
            f"取得設定: daily_period={opts.daily_period}, hourly_interval={opts.hourly_interval}, "
            f"minute_interval={opts.minute_interval}"
        )

        # 株価データ取得
        data = get_multi_timeframe_data(
            ticker=ticker,
            daily_period=opts.daily_period,
            daily_count=opts.daily_count,
            hourly_period=opts.hourly_period,
            hourly_count=opts.hourly_count,
            minute_period=opts.minute_period,
            minute_count=opts.minute_count,
            hourly_interval=opts.hourly_interval,
            minute_interval=opts.minute_interval,
        )

        # 簡易出力
        for timeframe, df in data.items():
            count = len(df) if df is not None else 0
            print(f"{timeframe}: {count} rows")

        # LLMに投資判断を依頼
        try:
            decision = advise_from_data(
                ticker=ticker,
                multi_data=data,
                objective="短期デイトレード観点で簡潔に判断してください",
                prefer_backend="openai",  # LangChain未使用時はopenai直呼び
                # modelは環境変数 LLM_MODEL があればそちらを使用
            )
            print("\n--- AI 投資判断 ---")
            print(f"判断: {decision.category_ja()} ({decision.action})")
            print(f"確信度: {decision.conviction}/100")
            print(f"期間: {decision.timeframe or '未指定'}")
            print("理由:")
            print(decision.reasoning)
            if decision.risks:
                print("リスク:")
                for i, r in enumerate(decision.risks, 1):
                    print(f"  {i}. {r}")
            if decision.targets:
                print("目標価格:")
                for k, v in decision.targets.items():
                    print(f"  {k}: {v}")

            # Excel用の行データを作成
            row = {
                "code": stock.code,
                "name": stock.name,
                "market": stock.market,
                "ticker": ticker,
                "category": decision.category_ja(),
                "action": decision.action,
                "conviction": decision.conviction,
                "timeframe": decision.timeframe or "",
                "reasoning": decision.reasoning,
                "risks": " | ".join(decision.risks) if decision.risks else "",
            }
            # targetsを平坦化して列化
            for k, v in (decision.targets or {}).items():
                row[f"target_{k}"] = v

            return row

        except Exception as e:
            print(f"AI分析に失敗しました: {e}")
            return None

    # 並列処理（MAX_WORKERS環境変数で同時実行数を調整。デフォルト4）
    max_workers = min(max(1, int(os.getenv("MAX_WORKERS", "4"))), len(ranking))
    print(f"\n並列実行: max_workers={max_workers}")
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {executor.submit(process_one, stock): stock for stock in ranking}
        for fut in as_completed(future_map):
            row = fut.result()
            if row:
                rows.append(row)

    # 集計結果をExcel出力（openpyxl未導入ならCSVにフォールバック）
    if rows:
        df = pd.DataFrame(rows)
        output_dir = "data"
        os.makedirs(output_dir, exist_ok=True)
        date_str = datetime.now().strftime("%Y%m%d")
        xlsx_path = os.path.join(output_dir, f"data_{date_str}.xlsx")
        try:
            # まずExcelに書き出し
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
                df.to_excel(writer, index=False, sheet_name="AI Decisions")

            # 書き出し後に列幅最適化・色付け
            try:
                from openpyxl import load_workbook
                from openpyxl.styles import PatternFill, Alignment, Font
                from openpyxl.utils import get_column_letter

                wb = load_workbook(xlsx_path)
                ws = wb["AI Decisions"] if "AI Decisions" in wb.sheetnames else wb.active

                # ヘッダー装飾
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill

                # オートフィルターとフリーズペイン
                ws.auto_filter.ref = ws.dimensions
                ws.freeze_panes = "A2"

                # 列名のインデックス取得
                headers = [c.value for c in ws[1]]
                col_index = {name: idx + 1 for idx, name in enumerate(headers)}

                # カテゴリ色分け（日本語カテゴリ）
                category_colors = {
                    "強い買い": "C6EFCE",  # 薄緑
                    "買い": "E2F0D9",      # さらに薄い緑
                    "ホールド": "F2F2F2",  # 薄灰
                    "売り": "FCE4D6",      # 薄赤
                    "強い売り": "F8CBAD",  # 濃い目の薄赤
                }
                cat_col = col_index.get("category")

                # 折返し対象列
                wrap_cols = [col_index.get("reasoning"), col_index.get("risks")]
                wrap_cols = [c for c in wrap_cols if c is not None]
                wrap_alignment = Alignment(wrap_text=True, vertical="top")

                # 列幅自動調整のために各列の最大文字数を計測
                max_len = {i: 0 for i in range(1, ws.max_column + 1)}
                for r in range(1, ws.max_row + 1):
                    for c in range(1, ws.max_column + 1):
                        val = ws.cell(row=r, column=c).value
                        s = "" if val is None else str(val)
                        # 改行は表示幅に影響するので一旦長さだけ計測
                        length = max([len(line) for line in s.splitlines()]) if "\n" in s else len(s)
                        if length > max_len[c]:
                            max_len[c] = length

                        # データ行のみ装飾
                        if r >= 2:
                            # カテゴリ色塗り
                            if cat_col and c == cat_col:
                                cat_text = str(val) if val is not None else ""
                                fill_color = category_colors.get(cat_text)
                                if fill_color:
                                    ws.cell(row=r, column=c).fill = PatternFill(
                                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                                    )
                            # 折返し列
                            if c in wrap_cols:
                                ws.cell(row=r, column=c).alignment = wrap_alignment

                # 列幅設定（余白分を加味）
                for c in range(1, ws.max_column + 1):
                    col_letter = get_column_letter(c)
                    # おおよその計算: 1文字 ≒ 1単位幅、余白+2、上限は適度に
                    width = min(max_len[c] + 2, 80)
                    ws.column_dimensions[col_letter].width = max(10, width)

                wb.save(xlsx_path)
            except Exception as deco_e:
                print(f"Excel整形に失敗しました: {deco_e}")

            print(f"\nExcelに保存しました: {xlsx_path}")
        except ImportError:
            csv_path = os.path.join(output_dir, f"data_{date_str}.csv")
            df.to_csv(csv_path, index=False, encoding="utf-8-sig")
            print(f"\nopenpyxl未導入のためCSVに保存しました: {csv_path}")


if __name__ == "__main__":
    main()